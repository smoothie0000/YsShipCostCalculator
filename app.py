import os
from datetime import datetime
from tkinter import filedialog
import customtkinter as ctk
from openpyxl import load_workbook, Workbook

import constants
import volume_strategy
import shipping_strategy

class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        ctk.set_appearance_mode("light")
        ctk.set_window_scaling(1.2)

        self.title("杨神快递费计算器")
        self.geometry("1000x800")
        self.wm_attributes("-topmost", True)

        self.selected_file = None
        self.product_rows = []
        self.next_row = 3

        self.volume_strategies = {
            "网格川字": volume_strategy.WangGeChuanZiStrategy(),
            "网格九脚": volume_strategy.WangGeJiuJiaoStrategy(),
            "平板四脚": volume_strategy.PingBanSiJiaoStrategy(),
            "平板九脚": volume_strategy.PingBanJiuJiaoStrategy(),
            "圆孔垫板": volume_strategy.YuanKongDianBanStrategy(),
        }

        self.shipping_strategies = {
            "德邦": shipping_strategy.DebangShippingStrategy(),
            "极兔": shipping_strategy.JituShippingStrategy(),
            "韵达": shipping_strategy.YundaShippingStrategy(),
            "中通": shipping_strategy.ZhongtongShippingStrategy(),
            "邮政": shipping_strategy.YouzhengShippingStrategy(),
            "快运（顺心捷达/壹米滴答）": shipping_strategy.KuaiyunShippingStrategy(),
        }

        self._build_ui()

    def _build_ui(self):
        # 标题
        title_label = ctk.CTkLabel(self, text="杨神快递费计算器", font=ctk.CTkFont(size=22, weight="bold"))
        title_label.grid(row=0, column=0, columnspan=4, pady=(20, 10))

        # 主frame
        self.frame = ctk.CTkFrame(self)
        self.frame.grid(row=1, column=0, padx=20, pady=10, sticky="nsew")

        # 快递省份选择
        province_label = ctk.CTkLabel(self.frame, text="快递省份", font=ctk.CTkFont(size=16))
        province_label.grid(row=0, column=0, padx=10, pady=0, sticky="w")

        self.province_combobox = ctk.CTkComboBox(self.frame, values=constants.china_provinces, width=200)
        self.province_combobox.set("上海市")
        self.province_combobox.grid(row=1, column=0, padx=10, pady=(0, 10), sticky="w")

        # 产品表头
        headers = ["选择产品", "数量", "单个产品尺寸(cm*cm*cm)", "单个产品重量(kg)", "产品总体积(cm³)"]
        for col, header in enumerate(headers):
            label = ctk.CTkLabel(self.frame, text=header, font=ctk.CTkFont(size=16))
            label.grid(row=2, column=col, padx=10, pady=(10, 0), sticky="w")

        self.add_product_row()

        # 添加产品按钮
        add_btn = ctk.CTkButton(self.frame, text="添加新产品", font=ctk.CTkFont(size=16), command=self.add_product_row)
        add_btn.grid(row=998, column=0, padx=10, pady=(10, 10), sticky="w")

        # 快递费用显示标签
        self.shipping_cost_label = ctk.CTkLabel(self.frame, text="快递费用:", font=ctk.CTkFont(size=16), anchor="w", justify="left")
        self.shipping_cost_label.grid(row=999, column=0, padx=10, pady=(35, 10), sticky="w")

        # 运费计算按钮
        calc_btn = ctk.CTkButton(self.frame, text="运费计算", font=ctk.CTkFont(size=16), command=self.calculate_shipping_cost)
        calc_btn.grid(row=1000, column=0, padx=10, pady=(0, 10), sticky="w")

        # 选择快递录入excel部分
        delivery_label = ctk.CTkLabel(self.frame, text="选择快递录入", font=ctk.CTkFont(size=24), anchor="w", justify="left")
        delivery_label.grid(row=1002, column=0, padx=10, pady=(35, 10), sticky="w")

        self.delivery_combobox = ctk.CTkComboBox(self.frame, values=list(self.shipping_strategies.keys()), width=150)
        self.delivery_combobox.set("德邦")
        self.delivery_combobox.grid(row=1003, column=0, padx=10, pady=(0, 10), sticky="w")

        excel_path_label = ctk.CTkLabel(self.frame, text="Excel文件路径:", font=ctk.CTkFont(size=16), anchor="w", justify="left")
        excel_path_label.grid(row=1004, column=0, padx=10, pady=(5, 10), sticky="w")

        self.file_path_entry = ctk.CTkEntry(self.frame, width=550)
        self.file_path_entry.grid(row=1005, column=0, columnspan=3, padx=10, pady=(0, 10), sticky="w")

        select_file_btn = ctk.CTkButton(self.frame, font=ctk.CTkFont(size=16), text="选择文件", command=self.select_files)
        select_file_btn.grid(row=1005, column=3, padx=10, pady=(0, 10), sticky="w")

        confirm_btn = ctk.CTkButton(self.frame, font=ctk.CTkFont(size=16), text="确定录入", command=self.confirm_entry)
        confirm_btn.grid(row=1006, column=0, padx=10, pady=(0, 10), sticky="w")

        self.result_label = ctk.CTkLabel(self.frame, text="", font=ctk.CTkFont(size=16), anchor="w", justify="left")
        self.result_label.grid(row=1007, column=0, columnspan=7, padx=10, pady=(5, 10), sticky="w")

    def add_product_row(self):
        row_widgets = {}

        # 产品选择
        product_cb = ctk.CTkComboBox(self.frame, values=constants.product_types, width=150)
        product_cb.set("网格川字")
        product_cb.grid(row=self.next_row, column=0, padx=10, pady=(0, 10), sticky="w")
        row_widgets['product'] = product_cb

        # 数量
        count_entry = ctk.CTkEntry(self.frame, placeholder_text="个", width=100)
        count_entry.grid(row=self.next_row, column=1, padx=10, pady=(0, 10), sticky="w")
        row_widgets['count'] = count_entry

        # 长宽高
        length_entry = ctk.CTkEntry(self.frame, placeholder_text="长", width=80)
        length_entry.grid(row=self.next_row, column=2, padx=(10, 2), pady=(0, 10), sticky="w")
        row_widgets['length'] = length_entry

        width_entry = ctk.CTkEntry(self.frame, placeholder_text="宽", width=80)
        width_entry.grid(row=self.next_row, column=2, padx=(90, 2), pady=(0, 10), sticky="w")
        row_widgets['width'] = width_entry

        height_entry = ctk.CTkEntry(self.frame, placeholder_text="高", width=80)
        height_entry.grid(row=self.next_row, column=2, padx=(170, 2), pady=(0, 10), sticky="w")
        row_widgets['height'] = height_entry

        # 重量
        weight_entry = ctk.CTkEntry(self.frame, placeholder_text="kg", width=100)
        weight_entry.grid(row=self.next_row, column=3, padx=10, pady=(0, 10), sticky="w")
        row_widgets['weight'] = weight_entry

        # 体积显示label
        volume_label = ctk.CTkLabel(self.frame, text="0 cm³", font=ctk.CTkFont(size=16))
        volume_label.grid(row=self.next_row, column=4, padx=10, pady=(0, 10), sticky="w")
        row_widgets['volume'] = volume_label

        # 计算体积按钮
        calc_btn = ctk.CTkButton(self.frame, text="计算体积", font=ctk.CTkFont(size=16), command=lambda rw=row_widgets: self.calculate_volume(rw))
        calc_btn.grid(row=self.next_row, column=5, padx=(30, 10), pady=(0, 10))
        row_widgets['calc_btn'] = calc_btn

        # 删除按钮
        del_btn = ctk.CTkButton(self.frame, text="删除", font=ctk.CTkFont(size=16), command=lambda rw=row_widgets: self.delete_product_row(rw))
        del_btn.grid(row=self.next_row, column=6, padx=10, pady=(0, 10))
        row_widgets['del_btn'] = del_btn

        self.product_rows.append(row_widgets)
        self.next_row += 1

    def delete_product_row(self, row_widgets):
        for widget in row_widgets.values():
            if isinstance(widget, ctk.CTkBaseClass):
                widget.destroy()
        self.product_rows.remove(row_widgets)

    def calculate_volume(self, row_widgets):
        try:
            count = int(row_widgets['count'].get())
            length = float(row_widgets['length'].get())
            width = float(row_widgets['width'].get())
            height = float(row_widgets['height'].get())
            product_type = row_widgets['product'].get()
            total_volume, total_length, total_width, total_height = self.volume_strategies[product_type].calculate(count, length, width, height)
            row_widgets['volume'].configure(text=f"{total_length} * {total_width} * {total_height} = {total_volume} cm³")
            return total_volume, total_length, total_width, total_height
        except Exception as e:
            row_widgets['volume'].configure(text="输入有误")
            return 0, 0, 0, 0

    def calculate_shipping_cost(self):
        province = self.province_combobox.get()
        total_volume = 0

        for row_widgets in self.product_rows:
            vol, _, _, _ = self.calculate_volume(row_widgets)
            total_volume += vol

        if total_volume == 0:
            self.shipping_cost_label.configure(text="请先正确填写产品体积和数量", text_color="red")
            return

        costs = {}
        for name, strategy in self.shipping_strategies.items():
            try:
                cost = strategy.calculate(total_volume, province)
                costs[name] = cost
            except Exception as e:
                costs[name] = f"不支持此省份"

        # 找最便宜的
        valid_costs = {k: v for k, v in costs.items() if isinstance(v, (int, float))}
        if valid_costs:
            cheapest_name = min(valid_costs, key=valid_costs.get)
            cheapest_cost = valid_costs[cheapest_name]
        else:
            cheapest_name = None
            cheapest_cost = None

        text_lines = ["快递费用:"]
        for name in self.shipping_strategies.keys():
            cost = costs.get(name, "无数据")
            if isinstance(cost, (int, float)):
                text_lines.append(f"{name}: {cost} 元")
            else:
                text_lines.append(f"{name}: {cost}")
        if cheapest_name:
            text_lines.append(f"\n最便宜为: {cheapest_name} {cheapest_cost} 元")

        self.shipping_cost_label.configure(text="\n".join(text_lines), text_color="black")

    def select_files(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            self.selected_file = file_path
            self.file_path_entry.delete(0, ctk.END)
            self.file_path_entry.insert(0, file_path)

    def confirm_entry(self):
        if not self.selected_file:
            self.result_label.configure(text="请选择Excel文件路径", text_color="red")
            return

        try:
            if not os.path.exists(self.selected_file):
                wb = Workbook()
                ws = wb.active
            else:
                wb = load_workbook(self.selected_file)
                ws = wb.active

            if ws.max_row == 1 and all(cell.value is None for cell in ws[1]):
                ws.append(["录入时间", "快递", "快递价格(元)", "产品类型", "数量(个)", "长(cm)", "宽(cm)", "高(cm)", "总体积(cm³)"])

            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            delivery_name = self.delivery_combobox.get()
            strategy = self.shipping_strategies.get(delivery_name)
            if not strategy:
                self.result_label.configure(text="无法识别快递", text_color="red")
                return

            for row_widgets in self.product_rows:
                total_volume, length, width, height = self.calculate_volume(row_widgets)
                product_type = row_widgets['product'].get() + " (多个不同规格产品一起邮寄，请手动检查体积是否正确)"
                count = float(row_widgets['count'].get())
                cost = strategy.calculate(total_volume, self.province_combobox.get())

                ws.append([now, delivery_name, cost, product_type, count, length, width, height, total_volume])

            wb.save(self.selected_file)
            self.result_label.configure(text="成功录入快递信息！", text_color="green")
        except Exception as e:
            self.result_label.configure(text=f"录入失败: {e}", text_color="red")


if __name__ == "__main__":
    app = App()
    app.mainloop()
