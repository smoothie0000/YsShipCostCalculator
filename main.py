import os
import math
import customtkinter as ctk
import constants
from datetime import datetime
from tkinter import filedialog
from openpyxl import load_workbook, Workbook

class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        ctk.set_appearance_mode("light")
        ctk.set_window_scaling(1.2)

        self.title("杨神快递费计算器")
        self.geometry("1000x800")
        self.wm_attributes("-topmost", True)

        # 标题
        title_label = ctk.CTkLabel(self, text="杨神快递费计算器", font=ctk.CTkFont(size=22, weight="bold"))
        title_label.grid(row=0, column=0, columnspan=4, pady=(20, 10))

        # 主frame
        self.frame = ctk.CTkFrame(self)
        self.frame.grid(row=1, column=0, padx=20, pady=10, sticky="nsew")

        # 快递省份
        province_label = ctk.CTkLabel(self.frame, text="快递省份", font=ctk.CTkFont(size=16))
        province_label.grid(row=0, column=0, padx=10, pady=0, sticky="w")

        self.province_combobox = ctk.CTkComboBox(self.frame, values=constants.china_provinces, width=200)
        self.province_combobox.set("上海市")
        self.province_combobox.grid(row=1, column=0, padx=10, pady=(0, 10), sticky="w")

        # 小标题
        labels = ["选择产品", "数量", "单个产品尺寸(cm*cm*cm)", "单个产品重量(kg)", "产品总体积(cm³)"]
        for idx, text in enumerate(labels):
            label = ctk.CTkLabel(self.frame, text=text, font=ctk.CTkFont(size=16))
            label.grid(row=2, column=idx, padx=10, pady=(10, 0), sticky="w")

        self.product_rows = []
        self.next_row = 3

        self.add_product_row()

        # 按钮
        self.add_product_button = ctk.CTkButton(self.frame, text="添加新产品", font=ctk.CTkFont(size=16), command=self.add_product_row)
        self.add_product_button.grid(row=998, column=0, padx=10, pady=(10, 10), sticky="w")

        self.shipping_cost_label = ctk.CTkLabel(self.frame, text="快递费用:", font=ctk.CTkFont(size=16), anchor="w", justify="left")
        self.shipping_cost_label.grid(row=999, column=0, padx=10, pady=(35, 10), sticky="w")

        self.calculate_shipping_button = ctk.CTkButton(self.frame, text="运费计算", font=ctk.CTkFont(size=16), command=self.calculate_shipping_cost)
        self.calculate_shipping_button.grid(row=1000, column=0, padx=10, pady=(0, 10), sticky="w")

        # 录入excel
        self.select_deliveru_entry_label = ctk.CTkLabel(self.frame, text="选择快递录入", font=ctk.CTkFont(size=24), anchor="w", justify="left")
        self.select_deliveru_entry_label.grid(row=1002, column=0, padx=10, pady=(35, 10), sticky="w")

        self.delivery_combobox = ctk.CTkComboBox(self.frame, values=constants.delivery_types, width=150)
        self.delivery_combobox.set("德邦")
        self.delivery_combobox.grid(row=1003, column=0, padx=10, pady=(0, 10), sticky="w")

        self.excel_path_label = ctk.CTkLabel(self.frame, text="Excel文件路径:", font=ctk.CTkFont(size=16), anchor="w", justify="left")
        self.excel_path_label.grid(row=1004, column=0, padx=10, pady=(5, 10), sticky="w")

        self.file_path_entry = ctk.CTkEntry(self.frame, width=550)
        self.file_path_entry.grid(row=1005, column=0, columnspan=3, padx=10, pady=(0, 10), sticky="w")

        self.select_button = ctk.CTkButton(self.frame, font=ctk.CTkFont(size=16), text="选择文件",command=self.select_files)
        self.select_button.grid(row=1005, column=3, padx=10, pady=(0, 10), sticky="w")

        self.confirm_button = ctk.CTkButton(self.frame, font=ctk.CTkFont(size=16), text="确定录入",command=self.confirm_entry)
        self.confirm_button.grid(row=1006, column=0, padx=10, pady=(0, 10), sticky="w")

        self.result_label = ctk.CTkLabel(self.frame, text="", font=ctk.CTkFont(size=16), anchor="w", justify="left")
        self.result_label.grid(row=1007, column=0, padx=10, pady=(5, 10), sticky="w")

    def select_files(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            self.select_file = file_path
            self.file_path_entry.delete(0, ctk.END)
            self.file_path_entry.insert(0,file_path)

    def confirm_entry(self):
        if not self.select_file:
            self.result_label.configure(text="请选择Excel文件路径", text_color="red")
            return

        try:
            if not os.path.exists(self.select_file):
                wb = Workbook()
                ws = wb.active
            else:
                wb = load_workbook(self.select_file)
                ws = wb.active

            if ws.max_row == 1 and all([cell.value is None for cell in ws[1]]):
                ws.append(["录入时间", "快递", "快递价格(元)", "产品类型", "数量(个)", "长(cm)", "宽(cm)", "高(cm)", "总体积(cm³)"])

            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for product in self.product_rows:
                total_volume, total_length, total_width, total_height = self.calculate_volume(product)
                product_type = product['product'].get() + "(多个不同规格产品一起邮寄，请手动检查体积是否正确)"
                count = float(product['count'].get())
                cost = 0
                if self.delivery_combobox.get() == '德邦':
                    cost = self.calculate_debang_cost(total_volume)
                elif self.delivery_combobox.get() == '极兔':
                    cost = self.calculate_jitu_cost(total_volume)
                elif self.delivery_combobox.get() == '韵达':
                    cost = self.calculate_yunda_cost(total_volume)
                elif self.delivery_combobox.get() == '中通':
                    cost = self.calculate_zhongtong_cost(total_volume)
                elif self.delivery_combobox.get() == '邮政':
                    cost = self.calculate_youzheng_cost(total_volume)
                elif self.delivery_combobox.get() == '快运（顺心捷达/壹米滴答）':
                    cost = self.calculate_kuaiyun_cost(total_volume)
                else:
                    self.result_label.configure(text="无法识别快递", text_color="red")
                    return

                ws.append([now, self.delivery_combobox.get(), cost, product_type, count, total_length, total_width, total_height, total_volume])
            wb.save(self.select_file)
            self.result_label.configure(text="成功录入快递信息！", text_color="green")
        except Exception as e:
            self.result_label.configure(text=f"录入失败: {e}", text_color="red")

    def add_product_row(self):
        row_widgets = {}

        product_combobox = ctk.CTkComboBox(self.frame, values=constants.product_types, width=150)
        product_combobox.set("网格川字")
        product_combobox.grid(row=self.next_row, column=0, padx=10, pady=(0, 10), sticky="w")
        row_widgets['product'] = product_combobox

        count_entry = ctk.CTkEntry(self.frame, placeholder_text="个", width=100)
        count_entry.grid(row=self.next_row, column=1, padx=10, pady=(0, 10), sticky="w")
        row_widgets['count'] = count_entry

        length_entry = ctk.CTkEntry(self.frame, placeholder_text="长", width=80)
        length_entry.grid(row=self.next_row, column=2, padx=(10, 2), pady=(0, 10), sticky="w")
        row_widgets['length'] = length_entry

        width_entry = ctk.CTkEntry(self.frame, placeholder_text="宽", width=80)
        width_entry.grid(row=self.next_row, column=2, padx=(90, 2), pady=(0, 10), sticky="w")
        row_widgets['width'] = width_entry

        height_entry = ctk.CTkEntry(self.frame, placeholder_text="高", width=80)
        height_entry.grid(row=self.next_row, column=2, padx=(170, 2), pady=(0, 10), sticky="w")
        row_widgets['height'] = height_entry

        weight_entry = ctk.CTkEntry(self.frame, placeholder_text="kg", width=100)
        weight_entry.grid(row=self.next_row, column=3, padx=10, pady=(0, 10), sticky="w")
        row_widgets['weight'] = weight_entry

        size_label = ctk.CTkLabel(self.frame, text="0 cm³", font=ctk.CTkFont(size=16))
        size_label.grid(row=self.next_row, column=4, padx=10, pady=(0, 10), sticky="w")
        row_widgets['size'] = size_label

        calculate_button = ctk.CTkButton(self.frame, text="计算体积", font=ctk.CTkFont(size=16),
                                         command=lambda rw=row_widgets: self.calculate_volume(rw))
        calculate_button.grid(row=self.next_row, column=5, padx=(30, 10), pady=(0, 10))
        row_widgets['button'] = calculate_button

        delete_button = ctk.CTkButton(self.frame, text="删除", font=ctk.CTkFont(size=16),
                                      command=lambda rw=row_widgets: self.delete_product_row(rw))
        delete_button.grid(row=self.next_row, column=6, padx=10, pady=(0, 10))
        row_widgets['delete'] = delete_button

        self.product_rows.append(row_widgets)
        self.next_row += 1

    def delete_product_row(self, row_widgets):
        # 删除界面上的组件
        for widget in row_widgets.values():
            if isinstance(widget, ctk.CTkBaseClass):
                widget.destroy()
        # 从列表中删除
        self.product_rows.remove(row_widgets)

    def calculate_volume(self, widgets):
        try:
            count = float(widgets['count'].get())
            length = float(widgets['length'].get())
            width = float(widgets['width'].get())
            height = float(widgets['height'].get())

            total_length = 0
            total_width = 0
            total_height = 0

            reminder = count % 2
            divide = math.floor(count / 2)
            single_volume = reminder * (length * width * height)
            product_type = widgets['product'].get()

            if product_type == "网格川字":
                bi_volume = length * (width + 18) * (height + 4)
                if count == 1:
                    total_length = length
                    total_width = width
                    total_height = height
                elif reminder == 0:
                    total_length = length
                    total_width = width + 18
                    total_height = divide * (height + 4)
                else:
                    total_length = length
                    total_width = width + 18
                    total_height = divide * (height + 4) + height
                total_volume = int(total_length * total_width * total_height)
            elif product_type == "网格九脚":
                if count == 1:
                    total_length = length
                    total_width = width
                    total_height = height
                else:
                    total_length = length
                    total_width = width
                    total_height = 5 * (count - 1) + 14
                total_volume = int(total_length * total_width * total_height)
            elif product_type == "平板四脚" or product_type == "平板六脚":
                bi_volume = length * (width + 13) * (height + 3)
                if count == 1:
                    total_length = length
                    total_width = width
                    total_height = height
                elif reminder == 0:
                    total_length = length
                    total_width = width + 13
                    total_height = divide * (height + 3)
                else:
                    total_length = length
                    total_width = width + 13
                    total_height = divide * (height + 3) + height
                total_volume = int(total_length * total_width * total_height)
            elif product_type == "平板九脚":
                bi_volume = length * (width + 18) * (height + 3)
                if count == 1:
                    total_length = length
                    total_width = width
                    total_height = height
                elif reminder == 0:
                    total_length = length
                    total_width = width + 18
                    total_height = divide * (height + 3)
                else:
                    total_length = length
                    total_width = width + 18
                    total_height = divide * (height + 3) + height
                total_volume = int(total_length * total_width * total_height)
            elif product_type == "圆孔垫板":
                total_length = length
                total_width = width
                total_height = count * height
                total_volume = int(total_length * total_width * total_height)
            else:
                widgets['size'].configure(text="产品类型错误")
                return 0, 0, 0, 0

            widgets['size'].configure(text=f"{total_length} * {total_width} * {total_height} = {total_volume} cm³")
            return total_volume, total_length, total_width, total_height

        except ValueError:
            widgets['size'].configure(text="输入有误")
            return 0, 0, 0, 0

    def calculate_debang_cost(self, total_volume):
        first_3kg_cost = 0
        over_per_kg_cost = 0
        if self.province_combobox.get() in ["江苏省", "浙江省", "上海市"]:
            first_3kg_cost = 7
            over_per_kg_cost = 1
        elif self.province_combobox.get() in ["广东省", "安徽省", "山东省",  "北京市", "天津市", "河北省", "河南省", "湖北省", "湖南省", "江西省", "山西省", "福建省"]:
            first_3kg_cost = 8
            over_per_kg_cost = 2.5
        elif self.province_combobox.get() in ["广西壮族自治区", "海南省", "云南省", "贵州省", "四川省", "重庆市", "黑龙江省", "吉林省", "辽宁省"]:
            first_3kg_cost = 9
            over_per_kg_cost = 3
        elif self.province_combobox.get() in ["陕西省", "甘肃省", "宁夏回族自治区", "青海省", "内蒙古自治区"]:
            first_3kg_cost = 12
            over_per_kg_cost = 3
        
        ship_cost = 1 # 1 元保单费
        volume_weight = round(total_volume / 12000)
        if volume_weight <= 3:
            ship_cost += first_3kg_cost
        else:
            ship_cost += first_3kg_cost + over_per_kg_cost * (volume_weight - 3)

        return round(ship_cost, 1)

    def calculate_jitu_cost(self, total_volume):
        smaller_than_half_kg_cost = 0
        half_to_one_kg_cost = 0
        one_to_two_kg_cost = 0
        two_to_three_kg_cost = 0
        first_3kg_cost = 0
        over_per_kg_cost = 0
        additional_cost = 0
        if self.province_combobox.get() in ["江苏省", "浙江省", "上海市", "安徽省"]:
            smaller_than_half_kg_cost = 2.3
            half_to_one_kg_cost = 2.8
            one_to_two_kg_cost = 3.7
            two_to_three_kg_cost = 5
            first_3kg_cost = 5
            over_per_kg_cost = 1
            if self.province_combobox.get() == "上海市":
                additional_cost = 1
        elif self.province_combobox.get() in ["福建省", "广东省", "江西省", "山东省", "河南省", "河北省", "天津市", "北京市", "湖南省", "湖北省"]:
            smaller_than_half_kg_cost = 2.3
            half_to_one_kg_cost = 2.8
            one_to_two_kg_cost = 3.7
            two_to_three_kg_cost = 5
            first_3kg_cost = 5
            over_per_kg_cost = 2
            if self.province_combobox.get() == "北京市":
                additional_cost = 1
        elif self.province_combobox.get() in ["黑龙江省", "吉林省", "辽宁省", "云南省", "重庆市", "广西壮族自治区", "贵州省", "四川省", "山西省", "陕西省"]:
            smaller_than_half_kg_cost = 2.3
            half_to_one_kg_cost = 2.8
            one_to_two_kg_cost = 3.7
            two_to_three_kg_cost = 5
            first_3kg_cost = 5
            over_per_kg_cost = 3
        elif self.province_combobox.get() in ["内蒙古自治区", "宁夏回族自治区", "青海省", "甘肃省", "海南省"]:
            smaller_than_half_kg_cost = 3.5
            half_to_one_kg_cost = 4
            one_to_two_kg_cost = 8
            two_to_three_kg_cost = 12
            first_3kg_cost = 12
            over_per_kg_cost = 4

        volume_weight = total_volume / 8000
        ship_cost = additional_cost * volume_weight
        if volume_weight <= 0.5:
            ship_cost += smaller_than_half_kg_cost
        elif volume_weight > 0.5 and volume_weight <= 1:
            ship_cost += half_to_one_kg_cost
        elif volume_weight > 1 and volume_weight <= 2:
            ship_cost += one_to_two_kg_cost
        elif volume_weight > 2 and volume_weight <= 3:
            ship_cost += two_to_three_kg_cost
        else:
            volume_weight = math.ceil(volume_weight)
            ship_cost += ship_cost + first_3kg_cost + over_per_kg_cost * (volume_weight - 3)
        return round(ship_cost, 1)

    def calculate_yunda_cost(self, total_volume):
        smaller_than_half_kg_cost = 0
        half_to_one_kg_cost = 0
        one_to_two_kg_cost = 0
        two_to_three_kg_cost = 0
        over_per_kg_cost = 0
        first_3kg_cost = 0
        if self.province_combobox.get() in ["江苏省", "浙江省", "安徽省"]:
            smaller_than_half_kg_cost = 2.5
            half_to_one_kg_cost = 2.8
            one_to_two_kg_cost = 3.6
            two_to_three_kg_cost = 4.6
            first_3kg_cost = 4.6
            over_per_kg_cost = 1
        elif self.province_combobox.get() in ["河北省", "天津市", "河南省", "湖南省", "湖北省", "山东省", "广东省", "江西省", "福建省"]:
            smaller_than_half_kg_cost = 2.5
            half_to_one_kg_cost = 2.8
            one_to_two_kg_cost = 3.6
            two_to_three_kg_cost = 4.6
            first_3kg_cost = 4.6
            over_per_kg_cost = 2
        elif self.province_combobox.get() in ["山西省", "陕西省", "广西壮族自治区", "四川省", "重庆市", "贵州省", "云南省", "黑龙江省", "辽宁省", "吉林省"]:
            smaller_than_half_kg_cost = 2.5
            half_to_one_kg_cost = 2.8
            one_to_two_kg_cost = 3.6
            two_to_three_kg_cost = 4.6
            first_3kg_cost = 4.6
            over_per_kg_cost = 2.8
        elif self.province_combobox.get() in ["内蒙古自治区", "甘肃省", "青海省", "宁夏回族自治区", "海南省"]:
            smaller_than_half_kg_cost = 3.5
            half_to_one_kg_cost = 3.8
            one_to_two_kg_cost = 4.8
            two_to_three_kg_cost = 5.8
            first_3kg_cost = 5.8
            over_per_kg_cost = 3.8
        elif self.province_combobox.get() in ["北京市"]:
            smaller_than_half_kg_cost = 3
            half_to_one_kg_cost = 3.8
            one_to_two_kg_cost = 4.8
            two_to_three_kg_cost = 5.8
            first_3kg_cost = 5.8
            over_per_kg_cost = 3
        elif self.province_combobox.get() in ["上海市"]:
            smaller_than_half_kg_cost = 3
            half_to_one_kg_cost = 3.8
            one_to_two_kg_cost = 4.8
            two_to_three_kg_cost = 5.8
            first_3kg_cost = 5.8
            over_per_kg_cost = 1.5

        ship_cost = 0
        volume_weight = total_volume / 8000
        if volume_weight <= 0.5:
            ship_cost = smaller_than_half_kg_cost
        elif volume_weight > 0.5 and volume_weight <= 1:
            ship_cost = half_to_one_kg_cost
        elif volume_weight > 1 and volume_weight <= 2:
            ship_cost = one_to_two_kg_cost
        elif volume_weight > 2 and volume_weight <= 3:
            ship_cost = two_to_three_kg_cost
        else:
            volume_weight = math.ceil(volume_weight)
            ship_cost = first_3kg_cost + over_per_kg_cost * (volume_weight - 3)
        return round(ship_cost, 1)
    
    def calculate_zhongtong_cost(self, total_volume):
        smaller_than_half_kg_cost = 0
        half_to_one_kg_cost = 0
        one_to_two_kg_cost = 0
        two_to_three_kg_cost = 0
        over_per_kg_cost = 0
        first_3kg_cost = 0
        additional_cost = 0
        calculate_with_volume = False
        price_per_square_meter = 0
        price_calculate_with_volume_fixed = 0
        if self.province_combobox.get() in ["江苏省", "浙江省", "安徽省"]:
            smaller_than_half_kg_cost = 2.3
            half_to_one_kg_cost = 2.8
            one_to_two_kg_cost = 3.5
            two_to_three_kg_cost = 5
            over_per_kg_cost = 1.5
            first_3kg_cost = 5
        elif self.province_combobox.get() in ["河北省", "天津市", "河南省", "湖南省", "湖北省", "山东省", "广东省", "江西省", "福建省"]:
            smaller_than_half_kg_cost = 2.3
            half_to_one_kg_cost = 2.8
            one_to_two_kg_cost = 3.5
            two_to_three_kg_cost = 5
            calculate_with_volume = True
            price_per_square_meter = 150
            price_calculate_with_volume_fixed = 5
        elif self.province_combobox.get() in ["山西省", "陕西省", "广西壮族自治区", "四川省", "重庆市", "贵州省", "云南省", "黑龙江省", "辽宁省", "吉林省"]:
            smaller_than_half_kg_cost = 2.3
            half_to_one_kg_cost = 2.8
            one_to_two_kg_cost = 3.5
            two_to_three_kg_cost = 5
            calculate_with_volume = True
            price_per_square_meter = 150
            price_calculate_with_volume_fixed = 5
        elif self.province_combobox.get() in ["内蒙古自治区", "甘肃省", "青海省", "宁夏回族自治区", "海南省"]:
            smaller_than_half_kg_cost = 3.5
            half_to_one_kg_cost = 4
            one_to_two_kg_cost = 5
            two_to_three_kg_cost = 6
            calculate_with_volume = True
            price_per_square_meter = 500
            price_calculate_with_volume_fixed = 0
        elif self.province_combobox.get() in ["北京市"]:
            smaller_than_half_kg_cost = 3
            half_to_one_kg_cost = 4
            one_to_two_kg_cost = 5
            two_to_three_kg_cost = 6.5
            additional_cost = 1.5
            calculate_with_volume = True
            price_per_square_meter = 150
            price_calculate_with_volume_fixed = 5.5
        elif self.province_combobox.get() in ["上海市"]:
            smaller_than_half_kg_cost = 3
            half_to_one_kg_cost = 4
            one_to_two_kg_cost = 5
            two_to_three_kg_cost = 5.5
            over_per_kg_cost = 1.5
            first_3kg_cost = 5.5
            additional_cost = 0.5

        ship_cost = additional_cost
        volume_weight = total_volume / 10000
        if volume_weight <= 0.5:
            ship_cost += smaller_than_half_kg_cost
        elif volume_weight > 0.5 and volume_weight <= 1:
            ship_cost += half_to_one_kg_cost
        elif volume_weight > 1 and volume_weight <= 2:
            ship_cost += one_to_two_kg_cost
        elif volume_weight > 2 and volume_weight <= 3:
            ship_cost += two_to_three_kg_cost
        elif volume_weight > 3 and calculate_with_volume:
            ship_cost += price_calculate_with_volume_fixed + price_per_square_meter * math.ceil(volume_weight) / 100
        else:
            ship_cost += first_3kg_cost + over_per_kg_cost * (math.ceil(volume_weight) - 3)

        return round(ship_cost, 1)

    def calculate_youzheng_cost(self, total_volume):
        first_kg_cost = 0
        over_per_kg_cost = 0
        if self.province_combobox.get() in ["上海市", "浙江省", "江苏省"]:
            first_kg_cost = 4
            over_per_kg_cost = 1
        if self.province_combobox.get() in ["安徽省"]:
            first_kg_cost = 4
            over_per_kg_cost = 1.2
        elif self.province_combobox.get() in ["北京市", "天津市", "河北省", "山西省", "山东省", "福建省", "江西省", "河南省", "湖北省", "湖南省", "广东省", ]:
            first_kg_cost = 5
            over_per_kg_cost = 1.7
        elif self.province_combobox.get() in ["广西壮族自治区", "陕西省", "重庆市", "贵州省"]:
            first_kg_cost = 5.5
            over_per_kg_cost = 1.7
        elif self.province_combobox.get() in ["辽宁省", "海南省", "宁夏回族自治区", "四川省", "甘肃省", "内蒙古自治区", "吉林省"]:
            first_kg_cost = 6
            over_per_kg_cost = 1.7
        elif self.province_combobox.get() in ["黑龙江省", "云南省"]:
            first_kg_cost = 6
            over_per_kg_cost = 2.6
        elif self.province_combobox.get() in ["青海省"]:
            first_kg_cost = 12
            over_per_kg_cost = 10
        
        ship_cost = 0
        volume_weight = math.ceil(total_volume / 12000)
        if volume_weight <= 1:
            ship_cost = first_kg_cost
        else:
            ship_cost = first_kg_cost + over_per_kg_cost * (volume_weight - 1)
        
        return round(ship_cost, 1)

    def calculate_kuaiyun_cost(self, total_volume):
        price_list = {
            "浙江省": [90, 25],
            "安徽省": [100, 30],
            "江苏省": [90, 25],
            "上海市": [100, 25],
            "福建省": [150, 35],
            "广东省": [150, 35],
            "江西省": [150, 35],
            "山东省": [150, 35],
            "北京市": [210, 50],
            "天津市": [160, 40],
            "河北省": [150, 45],
            "河南省": [150, 35],
            "湖北省": [150, 40],
            "湖南省": [150, 40],
            "重庆市": [240, 35],
            "四川省": [240, 50],
            "贵州省": [258, 40],
            "山西省": [210, 35],
            "陕西省": [210, 35],
            "广西壮族自治区": [218, 40],
            "辽宁省": [228, 40],
            "云南省": [258, 40],
            "黑龙江省": [228, 40],
            "吉林省": [228, 35],
            "甘肃省": [330, 50],
            "宁夏回族自治区": [330, 50],
            "海南省": [320, 50],
            "青海省": [330, 50],
            "内蒙古自治区": [330, 50],
            "西藏自治区": [-1, -1],
        }
        ship_cost = 0
        per_cube_meter_cost = price_list[self.province_combobox.get()][0]
        start_cost = price_list[self.province_combobox.get()][1]
        if per_cube_meter_cost * total_volume / 1000000 < start_cost:
            ship_cost = start_cost
        else:
            ship_cost = per_cube_meter_cost * total_volume / 1000000
        return round(ship_cost, 1)

    def calculate_shipping_cost(self):
        total_volume = 0
        # real_weight = 0
        for widgets in self.product_rows:
            _total_volume, total_length, total_width, total_height = self.calculate_volume(widgets)
            total_volume += _total_volume
            try:
                # weight = float(widgets['weight'].get())
                count = float(widgets['count'].get())
                # real_weight += weight * count
            except ValueError:
                self.shipping_cost_label.configure(text="产品数量不能为空")
                return

        debang_cost = self.calculate_debang_cost(total_volume)
        jitu_cost = self.calculate_jitu_cost(total_volume)
        yunda_cost = self.calculate_yunda_cost(total_volume)
        zhongtong_cost = self.calculate_zhongtong_cost(total_volume)
        kuaiyun_cost = self.calculate_kuaiyun_cost(total_volume)

        youzheng_fail = False
        for widgets in self.product_rows:
            _, total_length, total_width, total_height = self.calculate_volume(widgets)
            length = float(widgets['length'].get())
            width = float(widgets['width'].get())
            height = float(widgets['height'].get())
            if total_length > 100 or total_width > 100 or total_height > 100:
                youzheng_fail = True
                break
        if youzheng_fail:
            youzheng_cost = "单边长度不能超过100cm，产品不符合尺寸要求"
        else:
            youzheng_cost = self.calculate_youzheng_cost(total_volume)

        cheapest_name = "德邦"
        cheapest_cost = debang_cost
        if cheapest_cost > jitu_cost:
            cheapest_name = "极兔"
            cheapest_cost = jitu_cost
        if cheapest_cost > yunda_cost:
            cheapest_name = "韵达"
            cheapest_cost = yunda_cost
        if cheapest_cost > kuaiyun_cost:
            cheapest_name = "快运"
            cheapest_cost = kuaiyun_cost
        if cheapest_cost > zhongtong_cost:
            cheapest_name = "中通"
            cheapest_cost = zhongtong_cost
        if not youzheng_fail:
            if cheapest_cost > youzheng_cost:
                cheapest_name = "邮政"
                cheapest_cost = youzheng_cost

        if not youzheng_fail:
            self.shipping_cost_label.configure(
                text=(
                    f"快递费用:\n"
                    f"德邦: {debang_cost} 元\n"
                    f"极兔: {jitu_cost} 元\n"
                    f"韵达: {yunda_cost} 元\n"
                    f"中通: {zhongtong_cost} 元\n"
                    f"邮政: {youzheng_cost} 元\n"
                    f"快运(顺心捷达/壹米滴答): {kuaiyun_cost} 元\n\n"
                    f"最便宜为: {cheapest_name} {cheapest_cost} 元"
                )
            )
        else:
            self.shipping_cost_label.configure(
                text=(
                    f"快递费用:\n"
                    f"德邦: {debang_cost} 元\n"
                    f"极兔: {jitu_cost} 元\n"
                    f"韵达: {yunda_cost} 元\n"
                    f"中通: {zhongtong_cost} 元\n"
                    f"邮政: {youzheng_cost} \n"
                    f"快运(顺心捷达/壹米滴答): {kuaiyun_cost} 元\n\n"
                    f"最便宜为: {cheapest_name} {cheapest_cost} 元"
                )
            )


if __name__ == "__main__":
    app = App()
    app.mainloop()
